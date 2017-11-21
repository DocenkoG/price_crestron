# -*- coding: UTF-8 -*-
import xlrd                         # для .xls
import openpyxl                     # Для .xlsx
import re


def getCellXlsx(  row       # номер строки
                , col       # номер колонки 
                , isDigit   # Признак, числовое ли значение нужно из этого поля
                , sheet     #  лист XLSX
                ):
    '''
    Функция возвращает значение xls-ячейки в виде строки.    
    Для цифровых ячеек делается предварительное преобразование 
    в число (пустые и нечисловые значения преобразуются в "0")
    '''
    ccc = sheet.cell(row=row, column=col)
    cellType  = ccc.data_type
    cellValue = ccc.value
    if (isDigit == 'Y') : 
        if (cellValue == None) : 
            ss = '0'
        elif (cellType in ('n')) :                  # numeric
            if int(cellValue) == cellValue:
                ss = str(int(cellValue))
            else :
                ss = str(cellValue)
        else :
#           ss = '0'
            try:
                ss = str(float(cellValue.replace(',','.')))
            except ValueError as e:
                ss='0'
    else :
        if (cellValue == None) : 
            ss = ''
        elif (cellType in ('n')) :                    # numeric
            if int(cellValue) == cellValue:
                ss = str(int(cellValue))
            else :
                ss = str(cellValue)
        else :
            ss = str(cellValue)
    return ss



def getCell(  row       # номер строки
            , col       # номер колонки 
            , isDigit   # Признак, числовое ли значение нужно из этого поля
            , sheet     #  лист XLS
            ):
    '''
    Функция возвращает значение xls-ячейки в виде строки.    
    Для цифровых ячеек делается предварительное преобразование 
    в число (пустые и нечисловые значения преобразуются в "0")
    '''
    ccc = sheet.cell(row, col)
    cellType  = ccc.ctype
    cellValue = ccc.value
    if (isDigit == 'Y') : 
        if (cellValue == '') : 
            ss = '0'
        elif (cellType in (2,3)) :                  # numeric
            if int(cellValue) == cellValue:
                ss = str(int(cellValue))
            else :
                ss = str(cellValue)
        else :
            ss = str(float(cellValue))
            print(cellValue, ss)
    else :
        if (cellType in (2,3)) :                    # numeric
            if int(cellValue) == cellValue:
                ss = str(int(cellValue))
            else :
                ss = str(cellValue)
        else :
            ss = str(cellValue)
    return ss



def subInParentheses( sourceString):
    re_parentheses = re.compile('^.*\(([^)]*)\).*$', re.LOCALE | re.IGNORECASE )
    is_parentheses = re_parentheses.match(sourceString)
    if is_parentheses:                        # Файл соответствует шаблону имени
        key = is_parentheses.group(1)         # выделяю ключ из имени файла
    else:
        key = '' 
    return key



def currencyType(sheet, rowx, colx):
    '''
    Функция анализирует "формат ячейки" таблицы excel, является ли он "денежным"
    и какая валюта указана в этом формате.
    Распознаются не все валюты и способы их описания.
    '''
    c = sheet.cell(rowx, colx)
    xf = sheet.book.xf_list[c.xf_index]
    fmt_obj = sheet.book.format_map[xf.format_key]
    fmt_str = fmt_obj.format_str
    if '\u20bd' in fmt_str:
        val = 'RUB'
    elif '\xa3' in fmt_str:
        val = 'GBP'
    elif chr(8364) in fmt_str:
        val = 'EUR'
    elif (fmt_str.find('USD')>=0) or (fmt_str.find('[$$')>=0) :
        val = 'USD'
    else:
        val = ''
    return val

'''

[$$-409]#,##0.0
[$$-409]#,##0.0
[$$-409]#,##0.0
[$$-409]#,##0.0
[$$-409]#,##0.0
#,##0.0"р."
#,##0.0"р."
#,##0.0"р."
#,##0.0"р."
#,##0.0"р."
#
'''



def dump_cell(sheet, rowx, colx):
    c = sheet.cell(rowx, colx)
    xf = sheet.book.xf_list[c.xf_index]
    fmt_obj = sheet.book.format_map[xf.format_key]
    ccc = ord(fmt_obj.format_str[4])
    print( rowx, colx, repr(c.value), c.ctype, fmt_obj.type, ccc, chr(ccc) )
    #print( repr(fmt_obj.format_str))
    


def quoted(sss):
    if ((',' in sss) or ('"' in sss) or ('\n' in sss))  and not(sss[0]=='"' and sss[-1]=='"') :
        sss = '"'+sss.replace('"','""')+'"'
    return sss
