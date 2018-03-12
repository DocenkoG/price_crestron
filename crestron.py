# -*- coding: UTF-8 -*-
import os
import os.path
import logging
import logging.config
import sys
import configparser
import time
import shutil
import openpyxl                      # Для .xlsx
#import xlrd                          # для .xls
from   price_tools import getCellXlsx, getCell, quoted, dump_cell, currencyType, openX, sheetByName
import csv
import requests, lxml.html



def getXlsString(sh, i, in_columns_j):
    impValues = {}
    for item in in_columns_j.keys() :
        j = in_columns_j[item]-1
        if item in ('закупка','продажа','цена1') :
            if getCell(row=i, col=j, isDigit='N', sheet=sh).find('Звоните') >=0 :
                impValues[item] = '0.1'
            else :
                impValues[item] = getCell(row=i, col=j, isDigit='Y', sheet=sh)
            #print(sh, i, sh.cell( row=i, column=j).value, sh.cell(row=i, column=j).number_format, currencyType(sh, i, j))
        elif item == 'валюта_по_формату':
            impValues[item] = currencyType(row=i, col=j, sheet=sh)
        else:
            impValues[item] = getCell(row=i, col=j, isDigit='N', sheet=sh)
    return impValues



def getXlsxString(sh, i, in_columns_j):
    impValues = {}
    for item in in_columns_j.keys() :
        j = in_columns_j[item]
        if item in ('закупка','продажа','цена','цена1') :
            if getCellXlsx(row=i, col=j, isDigit='N', sheet=sh).find('Call for Pricing') >=0 :
                impValues[item] = '0.1'
            else :
                impValues[item] = getCellXlsx(row=i, col=j, isDigit='Y', sheet=sh)
            #print(sh, i, sh.cell( row=i, column=j).value, sh.cell(row=i, column=j).number_format, currencyType(sh, i, j))
        elif item == 'валюта_по_формату':
            impValues[item] = currencyType(row=i, col=j, sheet=sh)
        else:
            impValues[item] = getCellXlsx(row=i, col=j, isDigit='N', sheet=sh)
    return impValues



def convert_excel2csv(cfg):
    csvFName  = cfg.get('basic','filename_out')
    priceFName= cfg.get('basic','filename_in')
    sheetName = cfg.get('basic','sheetname')
    
    log.debug('Reading file ' + priceFName )
    sheet = sheetByName(fileName = priceFName, sheetName = sheetName)
    if not sheet :
        log.error("Нет листа "+sheetName+" в файле "+ priceFName)
        return False
    log.debug("Sheet   "+sheetName)
    out_cols = cfg.options("cols_out")
    in_cols  = cfg.options("cols_in")
    out_template = {}
    for vName in out_cols :
         out_template[vName] = cfg.get("cols_out", vName)
    in_cols_j = {}
    for vName in in_cols :
         in_cols_j[vName] = cfg.getint("cols_in",  vName)
    #brands,   discount     = config_read(cfgFName, 'discount')
    #for k in discount.keys():
    #    discount[k] = (100 - int(discount[k]))/100
    #print(discount)

    outFile = open( csvFName, 'w', newline='', encoding='CP1251', errors='replace')
    csvWriter = csv.DictWriter(outFile, fieldnames=out_cols )
    csvWriter.writeheader()

    '''                                     # Блок проверки свойств для распознавания групп      XLSX                                  
    for i in range(2393, 2397):                                                         
        i_last = i
        ccc = sheet.cell( row=i, column=in_cols_j['группа'] )
        print(i, ccc.value)
        print(ccc.font.name, ccc.font.sz, ccc.font.b, ccc.font.i, ccc.font.color.rgb, '------', ccc.fill.fgColor.rgb)
        print('------')
    '''
    '''                                     # Блок проверки свойств для распознавания групп      XLS                                  
    for i in range(0, 75):                                                         
        xfx = sheet.cell_xf_index(i, 0)
        xf  = book.xf_list[xfx]
        bgci  = xf.background.pattern_colour_index
        fonti = xf.font_index
        ccc = sheet.cell(i, 0)
        if ccc.value == None :
            print (i, colSGrp, 'Пусто!!!')
            continue
                                         # Атрибуты шрифта для настройки конфига
        font = book.font_list[fonti]
        print( '---------------------- Строка', i, '-----------------------', sheet.cell(i, 0).value)
        print( 'background_colour_index=',bgci)
        print( 'fonti=', fonti, '           xf.alignment.indent_level=', xf.alignment.indent_level)
        print( 'bold=', font.bold)
        print( 'weight=', font.weight)
        print( 'height=', font.height)
        print( 'italic=', font.italic)
        print( 'colour_index=', font.colour_index )
        print( 'name=', font.name)
    return
    '''

    recOut  ={}
    for i in range(1, sheet.max_row +1) :                                # xlsx
#   for i in range(1, sheet.nrows) :                                     # xls
        i_last = i
        try:
            impValues = getXlsxString(sheet, i, in_cols_j)              # xlsx
            #impValues = getXlsString(sheet, i, in_cols_j)                # xls
            #print( impValues )
            if impValues['цена1']=='0': # (ccc.value == None) or (ccc2.value == None) :     # Пустая строка
                pass
                #print( 'Пустая строка. i=',i, impValues )
            elif impValues['код_'] == '' or impValues['код_'] == 'Арт.' :  # Пустая строка
                print (i, 'Пусто!!!')
                continue
            else :                                                      # Обычная строка
                for outColName in out_template.keys() :
                    shablon = out_template[outColName]
                    for key in impValues.keys():
                        if shablon.find(key) >= 0 :
                            shablon = shablon.replace(key, impValues[key])
                    if (outColName == 'закупка') and ('*' in shablon) :
                        p = shablon.find("*")
                        vvv1 = float(shablon[:p])
                        vvv2 = float(shablon[p+1:])
                        shablon = str(round(vvv1 * vvv2, 2))
                    recOut[outColName] = shablon.strip()

#                if  recOut["валюта"]=="" :
#                    recOut["валюта"] = "USD"
                csvWriter.writerow(recOut)

        except Exception as e:
            print(e)
            if str(e) == "'NoneType' object has no attribute 'rgb'":
                pass
            else:
                log.debug('Exception: <' + str(e) + '> при обработке строки ' + str(i) +'.' )

    log.info('Обработано ' +str(i_last)+ ' строк.')
    outFile.close()



def download( cfg ):
    from selenium import webdriver
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.remote.remote_connection import LOGGER
    LOGGER.setLevel(logging.WARNING)
     
    retCode     = False
    filename_new= cfg.get('download','filename_new')
    filename_old= cfg.get('download','filename_old')
    login       = cfg.get('download','login'    )
    password    = cfg.get('download','password' )
    url_lk      = cfg.get('download','url_lk'   )
    url_file    = cfg.get('download','url_file' )

    download_path= os.path.join(os.getcwd(), 'tmp')
    if not os.path.exists(download_path):
        os.mkdir(download_path)

    for fName in os.listdir(download_path) :
        os.remove( os.path.join(download_path, fName))
    dir_befo_download = set(os.listdir(download_path))
        
    if os.path.exists('geckodriver.log') : os.remove('geckodriver.log')
    try:
        ffprofile = webdriver.FirefoxProfile()
        ffprofile.set_preference("browser.download.dir", download_path)
        ffprofile.set_preference("browser.download.folderList",2);
        ffprofile.set_preference("browser.helperApps.neverAsk.saveToDisk", 
                ",application/octet-stream" + 
                ",application/vnd.ms-excel" + 
                ",application/vnd.msexcel" + 
                ",application/x-excel" + 
                ",application/x-msexcel" + 
                ",application/zip" + 
                ",application/xls" + 
                ",application/vnd.ms-excel" +
                ",application/vnd.ms-excel.addin.macroenabled.12" +
                ",application/vnd.ms-excel.sheet.macroenabled.12" +
                ",application/vnd.ms-excel.template.macroenabled.12" +
                ",application/vnd.ms-excelsheet.binary.macroenabled.12" +
                ",application/vnd.ms-fontobject" +
                ",application/vnd.ms-htmlhelp" +
                ",application/vnd.ms-ims" +
                ",application/vnd.ms-lrm" +
                ",application/vnd.ms-officetheme" +
                ",application/vnd.ms-pki.seccat" +
                ",application/vnd.ms-pki.stl" +
                ",application/vnd.ms-word.document.macroenabled.12" +
                ",application/vnd.ms-word.template.macroenabed.12" +
                ",application/vnd.ms-works" +
                ",application/vnd.ms-wpl" +
                ",application/vnd.ms-xpsdocument" +
                ",application/vnd.openofficeorg.extension" +
                ",application/vnd.openxmformats-officedocument.wordprocessingml.document" +
                ",application/vnd.openxmlformats-officedocument.presentationml.presentation" +
                ",application/vnd.openxmlformats-officedocument.presentationml.slide" +
                ",application/vnd.openxmlformats-officedocument.presentationml.slideshw" +
                ",application/vnd.openxmlformats-officedocument.presentationml.template" +
                ",application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" +
                ",application/vnd.openxmlformats-officedocument.spreadsheetml.template" +
                ",application/vnd.openxmlformats-officedocument.wordprocessingml.template" +
                ",application/x-ms-application" +
                ",application/x-ms-wmd" +
                ",application/x-ms-wmz" +
                ",application/x-ms-xbap" +
                ",application/x-msaccess" +
                ",application/x-msbinder" +
                ",application/x-mscardfile" +
                ",application/x-msclip" +
                ",application/x-msdownload" +
                ",application/x-msmediaview" +
                ",application/x-msmetafile" +
                ",application/x-mspublisher" +
                ",application/x-msschedule" +
                ",application/x-msterminal" +
                ",application/x-mswrite" +
                ",application/xml" +
                ",application/xml-dtd" +
                ",application/xop+xml" +
                ",application/xslt+xml" +
                ",application/xspf+xml" +
                ",application/xv+xml" +
                ",application/excel")
        if os.name == 'posix':
            driver = webdriver.Firefox(ffprofile, executable_path=r'/usr/local/Cellar/geckodriver/0.19.1/bin/geckodriver')
        elif os.name == 'nt':
            driver = webdriver.Firefox(ffprofile)
        driver.implicitly_wait(30)
        
        driver.get(url_lk)
        time.sleep(2)
        driver.find_element_by_id("cntMain_ctrlLogin_loginControl_Username").clear()
        driver.find_element_by_id("cntMain_ctrlLogin_loginControl_Username").send_keys("mas@av-prom.ru")
        driver.find_element_by_id("cntMain_ctrlLogin_loginControl_Password").clear()
        driver.find_element_by_id("cntMain_ctrlLogin_loginControl_Password").send_keys("hsBNxPiRCY")
        driver.find_element_by_id("cntMain_ctrlLogin_loginControl_Login").click()
        time.sleep(3)
        driver.find_element_by_id("cntMain_btnDownload").click()
        time.sleep(3)
        driver.find_element_by_id("cntModal_chkTerms").click()
        time.sleep(3)
        driver.find_element_by_id("cntModal_btnOneSheetExcel").click()
        time.sleep(3)
        driver.find_element_by_id("cntModal_lbDownload").click()
        time.sleep(5)
        driver.find_element_by_xpath("(//button[@type='button'])[3]").click()
        time.sleep(3)
        driver.quit()

    except Exception as e:
        log.debug('Exception: <' + str(e) + '>')

    dir_afte_download = set(os.listdir(download_path))
    new_files = list( dir_afte_download.difference(dir_befo_download))
    print(new_files)
    if len(new_files) == 0 :        
        log.error( 'Не удалось скачать файл прайса ')
        return False
    elif len(new_files)>1 :
        log.error( 'Скачалось несколько файлов. Надо разбираться ...')
        return False
    else:   
        new_file = new_files[0]                                                     # загружен ровно один файл. 
        new_ext  = os.path.splitext(new_file)[-1].lower()
        DnewFile = os.path.join( download_path,new_file)
        new_file_date = os.path.getmtime(DnewFile)
        log.info( 'Скачанный файл ' +new_file + ' имеет дату ' + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(new_file_date) ) )
        
        print(new_ext)
        if new_ext in ('.xls','.xlsx','.xlsb','.xlsm','.csv'):
            if os.path.exists( filename_new) and os.path.exists( filename_old): 
                os.remove( filename_old)
                os.rename( filename_new, filename_old)
            if os.path.exists( filename_new) :
                os.rename( filename_new, filename_old)
            shutil.copy2( DnewFile, filename_new)
            return True

        elif new_ext == '.zip':  
            # ветка устаревшая, не проверялась                                      # Архив. Обработка не завершена
            log.debug( 'Zip-архив. Разархивируем.')
            work_dir = os.getcwd()                                                  
            os.chdir( os.path.join( download_path ))
            dir_befo_download = set(os.listdir(os.getcwd()))
            os.system('unzip -oj ' + new_file)
            os.remove(new_file)   
            dir_afte_download = set(os.listdir(os.getcwd()))
            new_files = list( dir_afte_download.difference(dir_befo_download))
            if len(new_files) == 1 :   
                new_file = new_files[0]                                             # разархивирован ровно один файл. 
                new_ext  = os.path.splitext(new_file)[-1]
                DnewFile = os.path.join( os.getcwd(),new_file)
                new_file_date = os.path.getmtime(DnewFile)
                log.debug( 'Файл из архива ' +DnewFile + ' имеет дату ' + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(new_file_date) )     )
                DnewPrice = DnewFile
            elif len(new_files) >1 :
                log.debug( 'В архиве не единственный файл. Надо разбираться.')
                DnewPrice = "dummy"
            else:
                log.debug( 'Нет новых файлов после разархивации. Загляни в папку юниттеста поставщика.')
                DnewPrice = "dummy"
            os.chdir(work_dir)




def config_read( cfgFName ):
    cfg = configparser.ConfigParser(inline_comment_prefixes=('#'))
    if  os.path.exists('private.cfg'):     
        cfg.read('private.cfg', encoding='utf-8')
    if  os.path.exists(cfgFName):     
        cfg.read( cfgFName, encoding='utf-8')
    else: 
        log.debug('Нет файла конфигурации '+cfgFName)
    return cfg



def is_file_fresh(fileName, qty_days):
    qty_seconds = qty_days *24*60*60 
    if os.path.exists( fileName):
        price_datetime = os.path.getmtime(fileName)
    else:
        log.error('Не найден файл  '+ fileName)
        return False

    if price_datetime+qty_seconds < time.time() :
        file_age = round((time.time()-price_datetime)/24/60/60)
        log.error('Файл "'+fileName+'" устарел!  Допустимый период '+ str(qty_days)+' дней, а ему ' + str(file_age) )
        return False
    else:
        return True



def make_loger():
    global log
    logging.config.fileConfig('logging.cfg')
    log = logging.getLogger('logFile')



def processing(cfgFName):
    log.info('----------------------- Processing '+cfgFName )
    cfg = config_read(cfgFName)
    filename_out = cfg.get('basic','filename_out')
    filename_in  = cfg.get('basic','filename_in')
    filename_new = cfg.get('download','filename_new')
    
    rc_download = False
    if cfg.has_section('download'):
        rc_download = download(cfg)
    if rc_download==True or is_file_fresh( filename_new, int(cfg.get('basic','срок годности'))):
        #os.system( 'marvel_converter_xlsx.xlsm')
        #convert_csv2csv(cfg)
        convert_excel2csv(cfg)
    folderName = os.path.basename(os.getcwd())
    if os.name == 'nt' :
        if os.path.exists(filename_out)  : shutil.copy2(filename_out , 'c://AV_PROM/prices/' + folderName +'/'+filename_out)
        if os.path.exists('python.log')  : shutil.copy2('python.log',  'c://AV_PROM/prices/' + folderName +'/python.log')
        if os.path.exists('python.log.1'): shutil.copy2('python.log.1','c://AV_PROM/prices/' + folderName +'/python.log.1')
    


def main( dealerName):
    """ Обработка прайсов выполняется согласно файлов конфигурации.
    Для этого в текущей папке должны быть файлы конфигурации, описывающие
    свойства файла и правила обработки. По одному конфигу на каждый 
    прайс или раздел прайса со своими правилами обработки
    """
    make_loger()
    log.info('          '+dealerName )
    for cfgFName in os.listdir("."):
        if cfgFName.startswith("cfg") and cfgFName.endswith(".cfg"):
            processing(cfgFName)


if __name__ == '__main__':
    myName = os.path.basename(os.path.splitext(sys.argv[0])[0])
    mydir    = os.path.dirname (sys.argv[0])
    print(mydir, myName)
    main( myName)
